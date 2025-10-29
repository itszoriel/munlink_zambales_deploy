"""Excel (XLSX) report utilities using openpyxl."""

from typing import List, Dict, Any, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def autosize(ws):
    for col in ws.columns:
        max_length = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(48, max_length + 2)


def generate_workbook(sheets: Dict[str, Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill(start_color='FFEEF7FF', end_color='FFEEF7FF', fill_type='solid')
    zebra_fill = PatternFill(start_color='FFF8FAFC', end_color='FFF8FAFC', fill_type='solid')
    bold = Font(bold=True)

    for name, spec in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        headers = [str(h) for h in (spec.get('headers', []) or [])]
        raw_rows = spec.get('rows', []) or []
        rows: List[List[Any]] = []
        for r in raw_rows:
            try:
                rows.append([("" if v is None else (str(v) if not isinstance(v, (int, float)) else v)) for v in r])
            except Exception:
                # Fallback to stringify whole row
                rows.append([str(v) for v in r])

        # Optional branded preheader
        municipality_name: Optional[str] = spec.get('municipality_name')
        report_title: Optional[str] = spec.get('title')
        gov_lines: Optional[list[str]] = spec.get('gov_lines')
        col_count = max(1, len(headers) or 1)
        # Center strictly over the table width
        span_cols = col_count
        last_row = 0
        if municipality_name or report_title or gov_lines:
            try:
                if municipality_name:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
                    c = ws.cell(row=1, column=1, value=municipality_name)
                    c.font = Font(bold=True, size=16)
                    # Center across whole merged span (set alignment on all cells in range for Excel rendering quirks)
                    for col in range(1, span_cols + 1):
                        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
                    last_row = 1
            except Exception:
                pass
            try:
                if report_title:
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
                    c = ws.cell(row=2, column=1, value=report_title)
                    c.font = Font(bold=True, size=12)
                    for col in range(1, span_cols + 1):
                        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center')
                    last_row = 2
            except Exception:
                pass
            try:
                if gov_lines:
                    r = last_row + 2 if last_row else 1
                    for i, line in enumerate(gov_lines):
                        start_col = col_count-1 if col_count>1 else 1
                        ws.merge_cells(start_row=r+i, start_column=start_col, end_row=r+i, end_column=col_count)
                        c = ws.cell(row=r+i, column=col_count, value=line)
                        c.font = Font(size=10)
                        c.alignment = Alignment(horizontal='right')
                    last_row = (r + len(gov_lines) - 1)
            except Exception:
                pass
            # Blank spacer row after header block
            last_row = (last_row or 0) + 1

        if headers:
            # Write header at the next row after preheader
            header_row_idx = (last_row + 1) if (municipality_name or report_title or gov_lines) else 1
            if header_row_idx > ws.max_row:
                # Ensure we are at correct row
                while ws.max_row < header_row_idx - 1:
                    ws.append([])
            ws.append(headers)
            for cell in ws[header_row_idx]:
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, r in enumerate(rows):
            ws.append(r)
            if idx % 2 == 1:
                for cell in ws[ws.max_row]:
                    cell.fill = zebra_fill

        autosize(ws)
        ws.freeze_panes = 'A2'

        # If there is an 'ID' column, left-align it and treat as text so it doesn't right-align by default
        try:
            if headers:
                try:
                    id_col_idx = headers.index('ID') + 1  # 1-based
                except ValueError:
                    id_col_idx = None
                if id_col_idx:
                    # Only apply left align to data rows, not the preheader or the header row
                    start_row = (header_row_idx + 1) if 'header_row_idx' in locals() else 3
                    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=id_col_idx, max_col=id_col_idx):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='left', vertical=cell.alignment.vertical if cell.alignment else 'center')
                            cell.number_format = '@'  # treat as text for left alignment consistency
        except Exception:
            pass

    return wb


def save_workbook(wb: Workbook, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


